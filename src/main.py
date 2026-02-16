import os
import re
import shutil
from datetime import datetime, date, timedelta
from typing import Optional, List
from functools import lru_cache

from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from sqlalchemy import func
from sqlalchemy.orm import joinedload
from openpyxl import load_workbook

from .database import get_session, init_db, Contractor, Employee, StopWord, Invoice, Act

from workalendar.europe import Russia

app = FastAPI()


@lru_cache(maxsize=1)
def _get_calendar() -> Russia:
    return Russia()


def get_russian_holidays(year: int) -> set:
    cal = _get_calendar()
    return {h[0] for h in cal.holidays(year)}


def is_weekend_or_holiday(d: date, holidays: set) -> bool:
    return d.weekday() >= 5 or d in holidays


def add_business_days(start_date: date, days: int, holidays: set) -> date:
    if days <= 0:
        return start_date

    current = start_date
    added = 0

    while added < days:
        current += timedelta(days=1)
        if not is_weekend_or_holiday(current, holidays):
            added += 1

    return current


STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
if os.path.exists(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

templates = Jinja2Templates(
    directory=os.path.join(os.path.dirname(__file__), "templates")
)

HTML_DIR = os.path.join(os.path.dirname(__file__), "templates")
if not os.path.exists(HTML_DIR):
    os.makedirs(HTML_DIR)


@app.on_event("startup")
def startup():
    init_db()


def normalize_contractor_name(name: str) -> str:
    if not name:
        return name
    name = name.strip()

    name = re.sub(r'["""\'\",;]', " ", name)
    name = re.sub(r"\s+", " ", name)

    name = re.sub(r"\s*\([^)]*\)\s*", " ", name)

    legal_forms = ["ООО", "ИП", "АО", "ЗАО", "ОАО", "ПАО", "НКО", "АНО", "ФГУП", "МУП"]
    pattern = r"(" + "|".join(re.escape(form) for form in legal_forms) + r")(?:\s|$)"
    match = re.search(pattern, name)
    if match:
        legal_form = match.group(1)
        name = name.replace(legal_form, "").strip() + " " + legal_form
    else:
        name = re.sub(r"\s+", " ", name).strip()

    return name.strip()


def parse_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return datetime(1899, 12, 30) + datetime.timedelta(days=int(value))
        except:
            return None
    if isinstance(value, str):
        value = value.strip()
        formats_with_time = [
            "%d.%m.%Y %H:%M",
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%y %H:%M",
            "%d.%m.%y %H:%M:%S",
            "%H:%M %d.%m.%Y",
            "%H:%M:%S %d.%m.%Y",
            "%H:%M %d.%m.%y",
            "%H:%M:%S %d.%m.%y",
            "%d.%m.%Y",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y/%m/%d",
        ]
        for fmt in formats_with_time:
            try:
                return datetime.strptime(value, fmt).date()
            except:
                continue
    return None


def parse_amount(value) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.replace(" ", "").replace(",", ".").strip()
        try:
            return float(value)
        except:
            return None
    return None


def get_or_create_contractor(session, name: str, inn: str = None) -> Contractor:
    normalized_name = normalize_contractor_name(name)
    contractor = (
        session.query(Contractor).filter(Contractor.name == normalized_name).first()
    )
    if not contractor:
        contractor = Contractor(name=normalized_name, inn=inn)
        session.add(contractor)
        session.flush()
    return contractor


def get_or_create_employee(session, full_name: str) -> Employee:
    if not full_name:
        return None
    parts = full_name.strip().split()
    first_name = parts[0] if len(parts) > 0 else ""
    last_name = parts[1] if len(parts) > 1 else ""
    middle_name = parts[2] if len(parts) > 2 else None

    employee = (
        session.query(Employee)
        .filter(Employee.last_name == last_name, Employee.first_name == first_name)
        .first()
    )

    if not employee:
        employee = Employee(
            last_name=last_name, first_name=first_name, middle_name=middle_name
        )
        session.add(employee)
        session.flush()

    return employee


def get_rpo_surnames() -> set:
    session = get_session()
    try:
        employees = session.query(Employee).all()
        return {e.last_name.lower() for e in employees}
    finally:
        session.close()


def check_employee_in_comment(comment: str, surnames: set) -> bool:
    if not comment:
        return False
    comment_lower = comment.lower()
    for surname in surnames:
        if surname in comment_lower:
            return True
    return False


@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request):
    return templates.TemplateResponse("dashboard.html", {"request": request})


@app.get("/unlinked-acts", response_class=HTMLResponse)
def unlinked_acts(request: Request):
    session = get_session()
    try:
        employees = session.query(Employee).all()
        contractors = session.query(Contractor).all()

        return templates.TemplateResponse(
            "unlinked_acts.html",
            {
                "request": request,
                "employees": employees,
                "contractors": contractors,
            },
        )
    finally:
        session.close()


@app.get("/linked-acts", response_class=HTMLResponse)
def linked_acts_page(request: Request):
    session = get_session()
    try:
        employees = session.query(Employee).all()
        contractors = session.query(Contractor).all()
        return templates.TemplateResponse(
            "linked_acts.html",
            {"request": request, "employees": employees, "contractors": contractors},
        )
    finally:
        session.close()


@app.get("/import", response_class=HTMLResponse)
def import_page(request: Request):
    session = get_session()
    try:
        stop_words = session.query(StopWord).all()
        return templates.TemplateResponse(
            "import.html", {"request": request, "stop_words": stop_words}
        )
    finally:
        session.close()


@app.post("/stop-words/add")
def add_stop_word(word: str = Form(...)):
    session = get_session()
    try:
        existing = session.query(StopWord).filter(StopWord.word == word).first()
        if not existing:
            sw = StopWord(word=word)
            session.add(sw)
            session.commit()
        return RedirectResponse("/import", status_code=303)
    finally:
        session.close()


@app.post("/stop-words/delete/{word_id}")
def delete_stop_word(word_id: int):
    session = get_session()
    try:
        sw = session.query(StopWord).filter(StopWord.id == word_id).first()
        if sw:
            session.delete(sw)
            session.commit()
        return RedirectResponse("/import", status_code=303)
    finally:
        session.close()


@app.post("/import-1c")
async def import_1c(file: UploadFile = File(...)):
    session = get_session()
    try:
        content = await file.read()
        temp_path = os.path.join(os.path.dirname(__file__), "temp_1c.xlsx")
        with open(temp_path, "wb") as f:
            f.write(content)

        wb = load_workbook(temp_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                col_map[h.strip()] = i + 1

        required_cols = [
            "№ п/п",
            "Дата",
            "Номер",
            "Сумма",
            "Контрагент",
            "Ответственный",
            "Комментарий",
            "Организация",
        ]
        for col in required_cols:
            if col not in col_map:
                return {"error": f"Missing column: {col}"}

        stop_words = [sw.word.lower() for sw in session.query(StopWord).all()]
        rpo_surnames = get_rpo_surnames()

        added = 0
        skipped_zero = 0
        skipped_delete = 0
        skipped_responsible = 0
        skipped_stopwords = 0
        skipped_duplicate = 0

        rows_detail = []

        for row in ws.iter_rows(min_row=2):
            try:
                number = str(row[col_map["Номер"] - 1].value or "").strip()
                invoice_date = parse_date(row[col_map["Дата"] - 1].value)
                amount = parse_amount(row[col_map["Сумма"] - 1].value)
                contractor_name = str(
                    row[col_map["Контрагент"] - 1].value or ""
                ).strip()
                responsible = str(row[col_map["Ответственный"] - 1].value or "").strip()
                responsible_parts = responsible.split()
                responsible_surname = (
                    responsible_parts[1] if len(responsible_parts) > 1 else ""
                )
                comment = str(row[col_map["Комментарий"] - 1].value or "").strip()
                comment_lower = comment.lower()
                org_group = str(row[col_map["Организация"] - 1].value or "").strip()

                row_info = {
                    "number": number,
                    "date": invoice_date.strftime("%d.%m.%Y") if invoice_date else "",
                    "amount": amount,
                    "contractor": contractor_name,
                    "responsible": responsible,
                    "comment": comment,
                    "status": "Импортирован",
                    "reasons": [],
                }

                if not amount or amount == 0:
                    row_info["status"] = "Пропущен"
                    row_info["reasons"].append("Сумма = 0 или пустая")
                    skipped_zero += 1

                if "удалить" in comment_lower or "заглушка" in comment_lower:
                    if row_info["status"] == "Импортирован":
                        row_info["status"] = "Пропущен"
                    row_info["reasons"].append(
                        "В комментарии есть 'удалить' или 'заглушка'"
                    )
                    skipped_delete += 1

                keep = False
                if responsible_surname.lower() in rpo_surnames:
                    keep = True
                    row_info["reasons"].append(
                        f"Ответственный '{responsible_surname}' найден в списке РПО"
                    )
                elif check_employee_in_comment(comment_lower, rpo_surnames):
                    keep = True
                    row_info["reasons"].append("Фамилия РПО найдена в комментарии")

                if not keep:
                    if row_info["status"] == "Импортирован":
                        row_info["status"] = "Пропущен"
                    row_info["reasons"].append(
                        f"Ответственный '{responsible_surname}' не относится к РПО/Продажи"
                    )
                    skipped_responsible += 1

                has_stop_word = False
                found_stop_words = []
                for sw in stop_words:
                    if sw in comment_lower:
                        has_stop_word = True
                        found_stop_words.append(sw)
                if has_stop_word:
                    if row_info["status"] == "Импортирован":
                        row_info["status"] = "Пропущен"
                    row_info["reasons"].append(
                        f"Найдены стоп-слова: {', '.join(found_stop_words)}"
                    )
                    skipped_stopwords += 1

                existing = (
                    session.query(Invoice)
                    .filter(
                        Invoice.number == number,
                        Invoice.date == invoice_date,
                        Invoice.amount == amount,
                    )
                    .first()
                )

                if existing:
                    if row_info["status"] == "Импортирован":
                        row_info["status"] = "Пропущен"
                    row_info["reasons"].append(
                        "Дубликат (счёт с такими реквизитами уже существует)"
                    )
                    skipped_duplicate += 1

                if row_info["status"] == "Импортирован":
                    contractor = get_or_create_contractor(session, contractor_name)

                    invoice = Invoice(
                        number=number,
                        date=invoice_date,
                        amount=amount,
                        contractor_id=contractor.id,
                        organization_group=org_group,
                        responsible_import=responsible,
                        comment=comment,
                        status="Не оплачен",
                    )
                    session.add(invoice)
                    added += 1

                rows_detail.append(row_info)

            except Exception as e:
                row_info = {
                    "number": "",
                    "date": "",
                    "amount": None,
                    "contractor": "",
                    "status": "Ошибка",
                    "reasons": [f"Ошибка обработки строки: {str(e)}"],
                }
                rows_detail.append(row_info)

        session.commit()
        os.remove(temp_path)

        return {
            "success": True,
            "added": added,
            "skipped_zero": skipped_zero,
            "skipped_delete": skipped_delete,
            "skipped_responsible": skipped_responsible,
            "skipped_stopwords": skipped_stopwords,
            "skipped_duplicate": skipped_duplicate,
            "rows_detail": rows_detail,
        }
    except Exception as e:
        session.rollback()
        return {"error": str(e)}
    finally:
        session.close()


@app.post("/import-sbis")
async def import_sbis(file: UploadFile = File(...)):
    session = get_session()
    try:
        content = await file.read()
        temp_path = os.path.join(os.path.dirname(__file__), "temp_sbis.xlsx")
        with open(temp_path, "wb") as f:
            f.write(content)

        wb = load_workbook(temp_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                col_map[h.strip()] = i + 1

        added = 0
        skipped_status = 0
        skipped_type = 0
        skipped_empty = 0
        skipped_duplicate = 0

        rows_detail = []

        for row in ws.iter_rows(min_row=2):
            try:
                doc_type = str(row[col_map["Тип документа"] - 1].value or "").strip()
                package_type = str(row[col_map["Тип пакета"] - 1].value or "").strip()
                status = str(row[col_map["Статус"] - 1].value or "").strip()
                amount = parse_amount(row[col_map["Сумма"] - 1].value)
                signing_date = parse_date(row[col_map["Завершено"] - 1].value)
                number = str(row[col_map["Номер"] - 1].value or "").strip()
                contractor_name = str(
                    row[col_map["Контрагент"] - 1].value or ""
                ).strip()
                inn_kpp = str(row[col_map["ИНН/КПП"] - 1].value or "").strip()
                inn = inn_kpp.split("/")[0] if inn_kpp else ""
                filename = str(row[col_map["Имя файла"] - 1].value or "").strip()

                row_info = {
                    "number": number,
                    "date": signing_date.strftime("%d.%m.%Y") if signing_date else "",
                    "amount": amount,
                    "contractor": contractor_name,
                    "inn": inn,
                    "filename": filename,
                    "doc_type": doc_type,
                    "package_type": package_type,
                    "status": status,
                    "import_status": "Импортирован",
                    "reasons": [],
                }

                if doc_type == "ЭДОСч" and package_type != "ДокОтгрИсх":
                    row_info["import_status"] = "Пропущен"
                    row_info["reasons"].append(f"Тип документа: {doc_type}")
                    skipped_type += 1

                if status != "Выполнение завершено успешно":
                    if row_info["import_status"] == "Импортирован":
                        row_info["import_status"] = "Пропущен"
                    row_info["reasons"].append(
                        f"Статус документа: '{status}' (ожидается 'Выполнение завершено успешно')"
                    )
                    skipped_status += 1

                if not amount or amount == 0:
                    if row_info["import_status"] == "Импортирован":
                        row_info["import_status"] = "Пропущен"
                    row_info["reasons"].append("Сумма = 0 или пустая")
                    skipped_empty += 1

                if not signing_date:
                    if row_info["import_status"] == "Импортирован":
                        row_info["import_status"] = "Пропущен"
                    row_info["reasons"].append("Дата подписания (Завершено) пустая")
                    skipped_empty += 1

                existing = (
                    session.query(Act)
                    .filter(
                        Act.number == number,
                        Act.signing_date == signing_date,
                        Act.amount == amount,
                    )
                    .first()
                )

                if existing:
                    if row_info["import_status"] == "Импортирован":
                        row_info["import_status"] = "Пропущен"
                    row_info["reasons"].append(
                        "Дубликат (акт с такими реквизитами уже существует)"
                    )
                    skipped_duplicate += 1

                if row_info["import_status"] == "Импортирован":
                    contractor = get_or_create_contractor(session, contractor_name, inn)

                    act = Act(
                        number=number,
                        filename=filename,
                        signing_date=signing_date,
                        amount=amount,
                        contractor_id=contractor.id,
                    )
                    session.add(act)
                    added += 1

                rows_detail.append(row_info)

            except Exception as e:
                row_info = {
                    "number": "",
                    "date": "",
                    "amount": None,
                    "contractor": "",
                    "inn": "",
                    "filename": "",
                    "doc_type": "",
                    "status": "",
                    "import_status": "Ошибка",
                    "reasons": [f"Ошибка обработки строки: {str(e)}"],
                }
                rows_detail.append(row_info)

        session.commit()
        os.remove(temp_path)

        return {
            "success": True,
            "added": added,
            "skipped_status": skipped_status,
            "skipped_type": skipped_type,
            "skipped_empty": skipped_empty,
            "skipped_duplicate": skipped_duplicate,
            "rows_detail": rows_detail,
        }
    except Exception as e:
        session.rollback()
        return {"error": str(e)}
    finally:
        session.close()


@app.post("/invoice/update/{invoice_id}")
def update_invoice(
    invoice_id: int,
    payment_date: Optional[str] = Form(None),
    deadline: Optional[str] = Form(None),
    motivated_person: Optional[str] = Form(None),
):
    session = get_session()
    try:
        invoice = session.query(Invoice).filter(Invoice.id == invoice_id).first()
        if invoice:
            if payment_date:
                invoice.payment_date = parse_date(payment_date)
            if deadline:
                invoice.deadline = parse_date(deadline)
            if motivated_person is not None:
                invoice.motivated_person = motivated_person
            session.commit()
        return {"success": True}
    finally:
        session.close()


@app.post("/act/update/{act_id}")
def update_act(
    act_id: int,
    responsible_manager: Optional[str] = Form(None),
    invoice_id: Optional[int] = Form(None),
):
    session = get_session()
    try:
        act = session.query(Act).filter(Act.id == act_id).first()
        if act:
            if responsible_manager is not None:
                act.responsible_manager = responsible_manager
            if invoice_id is not None:
                if invoice_id == 0:
                    act.invoice_id = None
                else:
                    act.invoice_id = invoice_id
            session.commit()
        return {"success": True}
    finally:
        session.close()


@app.post("/act/link/{act_id}")
def link_act(act_id: int, invoice_id: int = Form(...)):
    session = get_session()
    try:
        act = session.query(Act).filter(Act.id == act_id).first()
        if act:
            act.invoice_id = invoice_id
            session.commit()
        return RedirectResponse("/", status_code=303)
    finally:
        session.close()


@app.post("/act/unlink/{act_id}")
def unlink_act(act_id: int):
    session = get_session()
    try:
        act = session.query(Act).filter(Act.id == act_id).first()
        if act:
            act.invoice_id = None
            session.commit()
        return RedirectResponse("/", status_code=303)
    finally:
        session.close()


@app.post("/act/delete/{act_id}")
def delete_act(act_id: int):
    session = get_session()
    try:
        act = session.query(Act).filter(Act.id == act_id).first()
        if act:
            session.delete(act)
            session.commit()
            return {"success": True}
        return {"error": "Акт не найден", "success": False}
    except Exception as e:
        session.rollback()
        return {"error": str(e), "success": False}
    finally:
        session.close()


@app.post("/invoice/delete/{invoice_id}")
def delete_invoice(invoice_id: int):
    session = get_session()
    try:
        invoice = session.query(Invoice).filter(Invoice.id == invoice_id).first()
        if invoice:
            session.delete(invoice)
            session.commit()
            return {"success": True}
        return {"error": "Счёт не найден", "success": False}
    except Exception as e:
        session.rollback()
        return {"error": str(e), "success": False}
    finally:
        session.close()


@app.get("/employees", response_class=HTMLResponse)
def employees_page(request: Request):
    return templates.TemplateResponse("employees.html", {"request": request})


@app.get("/employees/list")
def list_employees():
    session = get_session()
    try:
        employees = session.query(Employee).all()
        return [
            {
                "id": e.id,
                "last_name": e.last_name,
                "first_name": e.first_name,
                "middle_name": e.middle_name,
                "department": e.department,
                "position": e.position,
            }
            for e in employees
        ]
    finally:
        session.close()


@app.post("/employees/add")
def add_employee(
    last_name: str = Form(...),
    first_name: str = Form(...),
    middle_name: Optional[str] = Form(None),
    department: Optional[str] = Form(None),
    position: Optional[str] = Form(None),
):
    session = get_session()
    try:
        existing = (
            session.query(Employee)
            .filter(
                Employee.last_name == last_name,
                Employee.first_name == first_name,
            )
            .first()
        )

        if existing:
            return {"error": "Сотрудник с такими ФИО уже существует", "success": False}

        employee = Employee(
            last_name=last_name,
            first_name=first_name,
            middle_name=middle_name,
            department=department,
            position=position,
        )
        session.add(employee)
        session.commit()
        return {"success": True}
    except Exception as e:
        session.rollback()
        return {"error": str(e), "success": False}
    finally:
        session.close()


@app.post("/employees/delete/{employee_id}")
def delete_employee(employee_id: int):
    session = get_session()
    try:
        employee = session.query(Employee).filter(Employee.id == employee_id).first()
        if employee:
            session.delete(employee)
            session.commit()
            return {"success": True}
        return {"error": "Сотрудник не найден", "success": False}
    except Exception as e:
        session.rollback()
        return {"error": str(e), "success": False}
    finally:
        session.close()


@app.post("/employees/update/{employee_id}")
def update_employee(
    employee_id: int,
    last_name: str = Form(...),
    first_name: str = Form(...),
    middle_name: Optional[str] = Form(None),
    department: Optional[str] = Form(None),
    position: Optional[str] = Form(None),
):
    session = get_session()
    try:
        employee = session.query(Employee).filter(Employee.id == employee_id).first()

        if not employee:
            return {"error": "Сотрудник не найден", "success": False}

        existing = (
            session.query(Employee)
            .filter(
                Employee.last_name == last_name,
                Employee.first_name == first_name,
                Employee.id != employee_id,
            )
            .first()
        )

        if existing:
            return {"error": "Сотрудник с такими ФИО уже существует", "success": False}

        employee.last_name = last_name
        employee.first_name = first_name
        employee.middle_name = middle_name
        employee.department = department
        employee.position = position

        session.commit()
        return {"success": True}
    except Exception as e:
        session.rollback()
        return {"error": str(e), "success": False}
    finally:
        session.close()


@app.get("/acts/free/{contractor_id}")
def get_free_acts(contractor_id: int):
    session = get_session()
    try:
        acts = (
            session.query(Act)
            .filter(Act.contractor_id == contractor_id, Act.invoice_id == None)
            .all()
        )
        return [
            {
                "id": a.id,
                "number": a.number,
                "signing_date": a.signing_date.strftime("%d.%m.%Y")
                if a.signing_date
                else "",
                "amount": a.amount,
                "responsible_manager": a.responsible_manager,
            }
            for a in acts
        ]
    finally:
        session.close()


@app.get("/acts/linked")
def get_linked_acts(
    contractor_id: Optional[str] = None,
    responsible_manager: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = "signing_date",
    sort_dir: Optional[str] = "desc",
):
    session = get_session()
    try:
        query = (
            session.query(Act)
            .filter(Act.invoice_id != None)
            .options(joinedload(Act.contractor), joinedload(Act.invoice))
        )

        if contractor_id and contractor_id.isdigit():
            query = query.filter(Act.contractor_id == int(contractor_id))

        if responsible_manager:
            query = query.filter(Act.responsible_manager == responsible_manager)

        if date_from:
            from_date = parse_date(date_from)
            if from_date:
                query = query.filter(Act.signing_date >= from_date)

        if date_to:
            to_date = parse_date(date_to)
            if to_date:
                query = query.filter(Act.signing_date <= to_date)

        sort_mapping = {
            "signing_date": Act.signing_date,
            "contractor_name": Contractor.name,
            "contractor_inn": Contractor.inn,
            "amount": Act.amount,
            "responsible_manager": Act.responsible_manager,
            "invoice_number": Invoice.number,
        }

        sort_column = sort_mapping.get(sort_by, Act.signing_date)

        if sort_by in ["contractor_name", "contractor_inn"]:
            query = query.join(Contractor, Act.contractor_id == Contractor.id)
        elif sort_by == "invoice_number":
            query = query.join(Invoice, Act.invoice_id == Invoice.id)

        if sort_dir == "desc":
            query = query.order_by(sort_column.desc())
        else:
            query = query.order_by(sort_column)

        acts = query.all()

        result = []
        for act in acts:
            contractor = act.contractor
            invoice = act.invoice

            result.append(
                {
                    "id": act.id,
                    "number": act.number,
                    "signing_date": act.signing_date.strftime("%d.%m.%Y")
                    if act.signing_date
                    else "",
                    "amount": act.amount,
                    "contractor_id": act.contractor_id,
                    "contractor_name": contractor.name if contractor else "",
                    "contractor_inn": contractor.inn if contractor else "",
                    "responsible_manager": act.responsible_manager,
                    "invoice_id": act.invoice_id,
                    "invoice_number": invoice.number if invoice else "",
                    "invoice_date": invoice.date.strftime("%d.%m.%Y")
                    if invoice and invoice.date
                    else "",
                }
            )

        return result
    finally:
        session.close()


@app.get("/acts/unlinked")
def get_unlinked_acts(
    contractor_id: Optional[str] = None,
    responsible_manager: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = "signing_date",
    sort_dir: Optional[str] = "desc",
):
    session = get_session()
    try:
        query = (
            session.query(Act)
            .filter(Act.invoice_id == None)
            .options(joinedload(Act.contractor))
        )

        if contractor_id and contractor_id.isdigit():
            query = query.filter(Act.contractor_id == int(contractor_id))

        if responsible_manager:
            query = query.filter(Act.responsible_manager == responsible_manager)

        if date_from:
            from_date = parse_date(date_from)
            if from_date:
                query = query.filter(Act.signing_date >= from_date)

        if date_to:
            to_date = parse_date(date_to)
            if to_date:
                query = query.filter(Act.signing_date <= to_date)

        sort_mapping = {
            "signing_date": Act.signing_date,
            "contractor_name": Contractor.name,
            "contractor_inn": Contractor.inn,
            "amount": Act.amount,
            "responsible_manager": Act.responsible_manager,
        }

        sort_column = sort_mapping.get(sort_by, Act.signing_date)

        if sort_by in ["contractor_name", "contractor_inn"]:
            query = query.join(Contractor, Act.contractor_id == Contractor.id)

        if sort_by == "has_available_invoices":
            acts = query.all()
            contractor_ids = [a.contractor_id for a in acts]
            available_invoices = (
                session.query(Invoice)
                .filter(
                    Invoice.contractor_id.in_(contractor_ids),
                    Invoice.status != "Оплачен",
                )
                .all()
            )
            invoices_by_contractor = {}
            for inv in available_invoices:
                if inv.contractor_id not in invoices_by_contractor:
                    invoices_by_contractor[inv.contractor_id] = []
                invoices_by_contractor[inv.contractor_id].append(inv)

            for act in acts:
                act._has_available = (
                    len(invoices_by_contractor.get(act.contractor_id, [])) > 0
                )

            acts.sort(
                key=lambda x: getattr(x, "_has_available", False),
                reverse=(sort_dir == "desc"),
            )
            result = []
            for act in acts:
                contractor = act.contractor

                result.append(
                    {
                        "id": act.id,
                        "number": act.number,
                        "signing_date": act.signing_date.strftime("%d.%m.%Y")
                        if act.signing_date
                        else "",
                        "amount": act.amount,
                        "contractor_id": act.contractor_id,
                        "contractor_name": contractor.name if contractor else "",
                        "contractor_inn": contractor.inn if contractor else "",
                        "responsible_manager": act.responsible_manager,
                    }
                )
        else:
            if sort_dir == "desc":
                query = query.order_by(sort_column.desc())
            else:
                query = query.order_by(sort_column)

            acts = query.all()

            result = []
            for act in acts:
                contractor = act.contractor

                result.append(
                    {
                        "id": act.id,
                        "number": act.number,
                        "signing_date": act.signing_date.strftime("%d.%m.%Y")
                        if act.signing_date
                        else "",
                        "amount": act.amount,
                        "contractor_id": act.contractor_id,
                        "contractor_name": contractor.name if contractor else "",
                        "contractor_inn": contractor.inn if contractor else "",
                        "responsible_manager": act.responsible_manager,
                    }
                )

        return result
    finally:
        session.close()


@app.get("/acts/by-invoice/{invoice_id}")
def get_acts_by_invoice(invoice_id: int):
    session = get_session()
    try:
        acts = session.query(Act).filter(Act.invoice_id == invoice_id).all()

        return [
            {
                "id": a.id,
                "number": a.number,
                "signing_date": a.signing_date.strftime("%d.%m.%Y")
                if a.signing_date
                else "",
                "amount": a.amount,
                "responsible_manager": a.responsible_manager,
            }
            for a in acts
        ]
    finally:
        session.close()


@app.post("/contractor/update-inn/{contractor_id}")
def update_contractor_inn(contractor_id: int, inn: str = Form(...)):
    session = get_session()
    try:
        contractor = (
            session.query(Contractor).filter(Contractor.id == contractor_id).first()
        )
        if contractor:
            contractor.inn = inn
            session.commit()
            return {"success": True}
        return {"error": "Контрагент не найден", "success": False}
    except Exception as e:
        session.rollback()
        return {"error": str(e), "success": False}
    finally:
        session.close()


@app.post("/invoice/calculate-deadline/{invoice_id}")
def calculate_deadline(invoice_id: int, days: int = Form(...)):
    session = get_session()
    try:
        invoice = session.query(Invoice).filter(Invoice.id == invoice_id).first()

        if not invoice:
            return {"error": "Счёт не найден", "success": False}

        if not invoice.payment_date:
            return {"error": "Не указана дата оплаты", "success": False}

        year = invoice.payment_date.year
        holidays = get_russian_holidays(year)

        deadline = add_business_days(invoice.payment_date, days, holidays)

        invoice.deadline = deadline
        invoice.deadline_days = days
        session.commit()

        return {"success": True, "deadline": deadline.strftime("%Y-%m-%d")}
    except Exception as e:
        session.rollback()
        return {"error": str(e), "success": False}
    finally:
        session.close()


@app.get("/invoices/list")
def list_invoices_filtered(
    contractor_id: Optional[int] = None,
    motivated_person: Optional[str] = None,
    payment_date_from: Optional[str] = None,
    payment_date_to: Optional[str] = None,
    sort_by: Optional[str] = "date",
    sort_dir: Optional[str] = "desc",
):
    session = get_session()
    try:
        from sqlalchemy import case

        query = session.query(Invoice)

        if contractor_id:
            query = query.filter(Invoice.contractor_id == contractor_id)

        if motivated_person:
            query = query.filter(Invoice.motivated_person == motivated_person)

        if payment_date_from:
            from_date = parse_date(payment_date_from)
            if from_date:
                query = query.filter(Invoice.payment_date >= from_date)

        if payment_date_to:
            to_date = parse_date(payment_date_to)
            if to_date:
                query = query.filter(Invoice.payment_date <= to_date)

        sort_mapping = {
            "date": Invoice.date,
            "deadline": Invoice.deadline,
            "contractor_name": Contractor.name,
            "contractor_inn": Contractor.inn,
            "responsible_import": Invoice.responsible_import,
            "motivated_person": Invoice.motivated_person,
            "payment_date": Invoice.payment_date,
            "acts_count": Invoice.id,
            "free_acts_count": Invoice.id,
        }

        sort_column = sort_mapping.get(sort_by, Invoice.deadline)

        if sort_by in ["contractor_name", "contractor_inn"]:
            query = query.join(Contractor, Invoice.contractor_id == Contractor.id)

        payment_date_nulls_last = case((Invoice.payment_date.is_(None), 1), else_=0)

        if sort_by in ["acts_count", "free_acts_count"]:
            invoices = query.options(joinedload(Invoice.contractor)).all()
        elif sort_dir == "desc":
            sort_column = sort_column.desc()
            query = query.order_by(payment_date_nulls_last.asc(), sort_column)
            invoices = query.options(joinedload(Invoice.contractor)).all()
        else:
            query = query.order_by(payment_date_nulls_last.asc(), sort_column)
            invoices = query.options(joinedload(Invoice.contractor)).all()

        invoice_ids = [inv.id for inv in invoices]
        contractor_ids = [inv.contractor_id for inv in invoices]

        all_linked_acts = (
            session.query(Act).filter(Act.invoice_id.in_(invoice_ids)).all()
            if invoice_ids
            else []
        )
        acts_by_invoice = {}
        for act in all_linked_acts:
            if act.invoice_id not in acts_by_invoice:
                acts_by_invoice[act.invoice_id] = []
            acts_by_invoice[act.invoice_id].append(act)

        all_free_acts = (
            session.query(Act)
            .filter(Act.contractor_id.in_(contractor_ids), Act.invoice_id == None)
            .all()
            if contractor_ids
            else []
        )
        free_acts_by_contractor = {}
        for act in all_free_acts:
            if act.contractor_id not in free_acts_by_contractor:
                free_acts_by_contractor[act.contractor_id] = []
            free_acts_by_contractor[act.contractor_id].append(act)

        result = []
        for inv in invoices:
            acts = acts_by_invoice.get(inv.id, [])
            sum_acts = sum(a.amount for a in acts)

            free_acts = free_acts_by_contractor.get(inv.contractor_id, [])
            free_acts_count = len(free_acts)

            contractor = inv.contractor

            result.append(
                {
                    "id": inv.id,
                    "number": inv.number,
                    "date": inv.date.strftime("%d.%m.%Y") if inv.date else "",
                    "amount": inv.amount,
                    "contractor_id": inv.contractor_id,
                    "contractor_name": contractor.name if contractor else "",
                    "contractor_inn": contractor.inn if contractor else "",
                    "payment_date": inv.payment_date.strftime("%Y-%m-%d")
                    if inv.payment_date
                    else "",
                    "deadline": inv.deadline.strftime("%Y-%m-%d")
                    if inv.deadline
                    else "",
                    "deadline_days": inv.deadline_days,
                    "responsible_import": inv.responsible_import,
                    "motivated_person": inv.motivated_person,
                    "status": inv.status,
                    "acts_count": len(acts),
                    "acts_sum": sum_acts,
                    "free_acts_count": free_acts_count,
                }
            )

        if sort_by == "acts_count":
            result.sort(key=lambda x: x["acts_count"], reverse=(sort_dir == "desc"))
        elif sort_by == "free_acts_count":
            result.sort(
                key=lambda x: x["free_acts_count"], reverse=(sort_dir == "desc")
            )

        return result
    finally:
        session.close()


@app.get("/contractors/list")
def list_contractors():
    session = get_session()
    try:
        contractors = session.query(Contractor).all()
        return [{"id": c.id, "name": c.name, "inn": c.inn} for c in contractors]
    finally:
        session.close()
